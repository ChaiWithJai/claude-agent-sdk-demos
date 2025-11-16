#!/usr/bin/env python3
"""
Quarterly Tax Calculator
Calculates estimated quarterly taxes for LLC owners and self-employed individuals.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_quarterly_tax_calculator(filename='Quarterly_Tax_Calculator.xlsx'):
    """Create a quarterly tax estimate calculator for self-employed individuals."""

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Q1 2024"

    # Title
    sheet['A1'] = 'QUARTERLY TAX ESTIMATE - Q1 2024'
    sheet['A1'].font = Font(bold=True, size=16)
    sheet.merge_cells('A1:D1')

    # Disclaimer
    sheet['A2'] = 'DISCLAIMER: This is for estimation purposes only. Consult a CPA for tax advice.'
    sheet['A2'].font = Font(italic=True, size=10, color='FF0000')
    sheet.merge_cells('A2:D2')

    # Column widths
    sheet.column_dimensions['A'].width = 35
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 30

    # Section 1: Income
    row = 4
    sheet[f'A{row}'] = 'INCOME (Q1: January - March)'
    sheet[f'A{row}'].font = Font(bold=True, size=12, color='FFFFFF')
    sheet[f'A{row}'].fill = PatternFill('solid', start_color='4472C4')
    sheet.merge_cells(f'A{row}:B{row}')

    row += 1
    sheet[f'A{row}'] = 'Gross Business Income (LLC)'
    sheet[f'B{row}'] = 25000
    sheet[f'B{row}'].number_format = '$#,##0.00'
    sheet[f'B{row}'].font = Font(color='0000FF')

    row += 1
    sheet[f'A{row}'] = '1099-NEC Income'
    sheet[f'B{row}'] = 5000
    sheet[f'B{row}'].number_format = '$#,##0.00'
    sheet[f'B{row}'].font = Font(color='0000FF')

    row += 1
    sheet[f'A{row}'] = 'Other Business Income'
    sheet[f'B{row}'] = 0
    sheet[f'B{row}'].number_format = '$#,##0.00'
    sheet[f'B{row}'].font = Font(color='0000FF')

    income_total_row = row + 1
    sheet[f'A{income_total_row}'] = 'Total Quarterly Income'
    sheet[f'B{income_total_row}'] = f'=SUM(B{row-2}:B{row})'
    sheet[f'A{income_total_row}'].font = Font(bold=True)
    sheet[f'B{income_total_row}'].font = Font(bold=True)
    sheet[f'B{income_total_row}'].number_format = '$#,##0.00'
    sheet[f'B{income_total_row}'].fill = PatternFill('solid', start_color='D9E1F2')

    # Section 2: Expenses
    row = income_total_row + 2
    sheet[f'A{row}'] = 'BUSINESS EXPENSES (Q1)'
    sheet[f'A{row}'].font = Font(bold=True, size=12, color='FFFFFF')
    sheet[f'A{row}'].fill = PatternFill('solid', start_color='C65911')
    sheet.merge_cells(f'A{row}:B{row}')

    expenses = [
        ('Advertising & Marketing', 500),
        ('Office Supplies', 200),
        ('Software & Subscriptions', 450),
        ('Professional Services', 1000),
        ('Business Meals (50% deductible)', 300),
        ('Travel', 800),
        ('Insurance', 600),
        ('Auto Expenses (or Mileage)', 750),
        ('Home Office', 900),
        ('Other Expenses', 200),
    ]

    expense_start = row + 1
    for expense_name, amount in expenses:
        row += 1
        sheet[f'A{row}'] = expense_name
        sheet[f'B{row}'] = amount
        sheet[f'B{row}'].number_format = '$#,##0.00'
        sheet[f'B{row}'].font = Font(color='0000FF')

    expense_total_row = row + 1
    sheet[f'A{expense_total_row}'] = 'Total Quarterly Expenses'
    sheet[f'B{expense_total_row}'] = f'=SUM(B{expense_start}:B{row})'
    sheet[f'A{expense_total_row}'].font = Font(bold=True)
    sheet[f'B{expense_total_row}'].font = Font(bold=True)
    sheet[f'B{expense_total_row}'].number_format = '$#,##0.00'
    sheet[f'B{expense_total_row}'].fill = PatternFill('solid', start_color='F4B084')

    # Section 3: Net Profit
    row = expense_total_row + 2
    sheet[f'A{row}'] = 'NET PROFIT (Income - Expenses)'
    sheet[f'B{row}'] = f'=B{income_total_row}-B{expense_total_row}'
    sheet[f'A{row}'].font = Font(bold=True, size=12)
    sheet[f'B{row}'].font = Font(bold=True, size=12, color='FFFFFF')
    sheet[f'B{row}'].fill = PatternFill('solid', start_color='70AD47')
    sheet[f'B{row}'].number_format = '$#,##0.00'

    net_profit_row = row

    # Section 4: Tax Calculations
    row += 2
    sheet[f'A{row}'] = 'TAX CALCULATIONS'
    sheet[f'A{row}'].font = Font(bold=True, size=12, color='FFFFFF')
    sheet[f'A{row}'].fill = PatternFill('solid', start_color='7030A0')
    sheet.merge_cells(f'A{row}:B{row}')

    # Self-Employment Tax
    row += 1
    sheet[f'A{row}'] = 'Self-Employment Tax Base (92.35%)'
    sheet[f'B{row}'] = f'=B{net_profit_row}*0.9235'
    sheet[f'B{row}'].number_format = '$#,##0.00'

    se_base_row = row

    row += 1
    sheet[f'A{row}'] = 'Self-Employment Tax (15.3%)'
    sheet[f'B{row}'] = f'=B{se_base_row}*0.153'
    sheet[f'B{row}'].number_format = '$#,##0.00'
    sheet[f'B{row}'].font = Font(bold=True)

    se_tax_row = row

    # Estimated Income Tax
    row += 2
    sheet[f'A{row}'] = 'Estimated Income Tax (22% bracket)'
    sheet[f'B{row}'] = f'=B{net_profit_row}*0.22'
    sheet[f'B{row}'].number_format = '$#,##0.00'
    sheet[f'B{row}'].font = Font(bold=True)
    sheet[f'C{row}'] = 'Adjust based on your tax bracket'
    sheet[f'C{row}'].font = Font(italic=True, size=9)

    income_tax_row = row

    # Total Quarterly Tax
    row += 2
    sheet[f'A{row}'] = 'TOTAL QUARTERLY TAX'
    sheet[f'B{row}'] = f'=B{se_tax_row}+B{income_tax_row}'
    sheet[f'A{row}'].font = Font(bold=True, size=14, color='FFFFFF')
    sheet[f'B{row}'].font = Font(bold=True, size=14, color='FFFFFF')
    sheet[f'A{row}'].fill = PatternFill('solid', start_color='7030A0')
    sheet[f'B{row}'].fill = PatternFill('solid', start_color='7030A0')
    sheet[f'B{row}'].number_format = '$#,##0.00'

    total_tax_row = row

    # Payment Info
    row += 2
    sheet[f'A{row}'] = 'QUARTERLY PAYMENT DUE'
    sheet[f'A{row}'].font = Font(bold=True, size=12)

    row += 1
    sheet[f'A{row}'] = 'Amount to Pay (1040-ES)'
    sheet[f'B{row}'] = f'=B{total_tax_row}'
    sheet[f'B{row}'].number_format = '$#,##0.00'
    sheet[f'B{row}'].font = Font(bold=True, size=12, color='FFFFFF')
    sheet[f'B{row}'].fill = PatternFill('solid', start_color='FF0000')

    row += 1
    sheet[f'A{row}'] = 'Q1 Due Date: April 15, 2024'
    sheet[f'A{row}'].font = Font(italic=True)

    row += 1
    sheet[f'A{row}'] = 'Payment Method: IRS Direct Pay, EFTPS, or check'
    sheet[f'A{row}'].font = Font(italic=True, size=9)

    # Annual Projection
    row += 3
    sheet[f'A{row}'] = 'ANNUAL PROJECTION (Q1 x 4)'
    sheet[f'A{row}'].font = Font(bold=True, size=12)
    sheet.merge_cells(f'A{row}:B{row}')

    row += 1
    sheet[f'A{row}'] = 'Projected Annual Income'
    sheet[f'B{row}'] = f'=B{income_total_row}*4'
    sheet[f'B{row}'].number_format = '$#,##0.00'

    row += 1
    sheet[f'A{row}'] = 'Projected Annual Expenses'
    sheet[f'B{row}'] = f'=B{expense_total_row}*4'
    sheet[f'B{row}'].number_format = '$#,##0.00'

    row += 1
    sheet[f'A{row}'] = 'Projected Annual Net Profit'
    sheet[f'B{row}'] = f'=B{net_profit_row}*4'
    sheet[f'B{row}'].number_format = '$#,##0.00'
    sheet[f'B{row}'].font = Font(bold=True)

    row += 1
    sheet[f'A{row}'] = 'Projected Annual Tax Liability'
    sheet[f'B{row}'] = f'=B{total_tax_row}*4'
    sheet[f'B{row}'].number_format = '$#,##0.00'
    sheet[f'B{row}'].font = Font(bold=True, color='FFFFFF')
    sheet[f'B{row}'].fill = PatternFill('solid', start_color='FF0000')

    # Instructions sheet
    instructions = wb.create_sheet('Instructions')
    instructions['A1'] = 'Quarterly Tax Calculator Instructions'
    instructions['A1'].font = Font(bold=True, size=14)

    instructions_text = [
        '',
        'How to Use This Calculator:',
        '',
        '1. Income Section:',
        '   - Enter your total business income for the quarter',
        '   - Include all LLC income, 1099-NEC payments, and other business revenue',
        '',
        '2. Expenses Section:',
        '   - Enter deductible business expenses for the quarter',
        '   - Use actual expenses from your expense tracker',
        '   - Remember: Business meals are 50% deductible',
        '',
        '3. Tax Calculations:',
        '   - Self-Employment Tax: 15.3% on 92.35% of net profit',
        '   - Income Tax: Estimated based on your tax bracket (adjust as needed)',
        '',
        '4. Quarterly Payment:',
        '   - Use Form 1040-ES to make quarterly payments',
        '   - Pay online at IRS.gov/payments or use EFTPS',
        '   - Keep records of all payments made',
        '',
        'Quarterly Due Dates:',
        '   - Q1 (Jan-Mar): April 15',
        '   - Q2 (Apr-May): June 15',
        '   - Q3 (Jun-Aug): September 15',
        '   - Q4 (Sep-Dec): January 15 (following year)',
        '',
        'Important Notes:',
        '   - This is an estimate only - actual tax may vary',
        '   - Adjust income tax rate based on your total income and filing status',
        '   - Consider state income tax if applicable',
        '   - Deduct 50% of self-employment tax from income',
        '   - Keep detailed records for tax filing',
        '',
        'Tax Brackets (2024 Single Filer):',
        '   - 10%: $0 to $11,600',
        '   - 12%: $11,601 to $47,150',
        '   - 22%: $47,151 to $100,525',
        '   - 24%: $100,526 to $191,950',
        '   - (Higher brackets for income above)',
        '',
        'DISCLAIMER:',
        '   This calculator provides estimates only and does not constitute',
        '   professional tax advice. Tax laws change annually. Consult with a',
        '   licensed CPA or tax professional for advice specific to your situation.',
        '',
        f'Created: {datetime.now().strftime("%Y-%m-%d %H:%M")}',
    ]

    for i, text in enumerate(instructions_text, start=2):
        instructions[f'A{i}'] = text
        if text.strip().startswith('-'):
            instructions[f'A{i}'].alignment = Alignment(indent=2)
        elif text.strip().startswith('•'):
            instructions[f'A{i}'].alignment = Alignment(indent=3)

    # Save workbook
    wb.save(filename)
    print(f"✓ Quarterly tax calculator created: {filename}")
    print(f"  - Q1 2024 template included")
    print(f"  - Self-employment tax calculation")
    print(f"  - Income tax estimate")
    print(f"  - Annual projection")
    print(f"  - Instructions sheet added")
    print(f"\nNOTE: Adjust income tax rate based on your total income and tax bracket")

if __name__ == '__main__':
    create_quarterly_tax_calculator()
