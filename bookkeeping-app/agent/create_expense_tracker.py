#!/usr/bin/env python3
"""
Expense Tracker Template Generator
Creates a spreadsheet for tracking business expenses with tax deduction calculations.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_expense_tracker(filename='Expense_Tracker.xlsx'):
    """Create an expense tracking spreadsheet with IRS categories."""

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Expenses"

    # Headers
    headers = ['Date', 'Vendor', 'Category', 'Description', 'Amount',
               'Business %', 'Deductible', 'Payment Method', 'Receipt']
    sheet.append(headers)

    # Header formatting
    header_fill = PatternFill('solid', start_color='C65911')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Column widths
    sheet.column_dimensions['A'].width = 12  # Date
    sheet.column_dimensions['B'].width = 20  # Vendor
    sheet.column_dimensions['C'].width = 20  # Category
    sheet.column_dimensions['D'].width = 30  # Description
    sheet.column_dimensions['E'].width = 12  # Amount
    sheet.column_dimensions['F'].width = 12  # Business %
    sheet.column_dimensions['G'].width = 12  # Deductible
    sheet.column_dimensions['H'].width = 15  # Payment Method
    sheet.column_dimensions['I'].width = 12  # Receipt

    # Sample data
    sample_data = [
        ['2024-01-05', 'Office Depot', 'Office Supplies', 'Printer paper and pens', 45.50, 1.00, '=E2*F2', 'Credit Card', 'Yes'],
        ['2024-01-12', 'AWS', 'Software/SaaS', 'Monthly cloud hosting', 125.00, 1.00, '=E3*F3', 'Credit Card', 'Yes'],
        ['2024-01-18', 'Coffee Shop', 'Meals - Business', 'Client meeting lunch', 78.00, 0.50, '=E4*F4', 'Credit Card', 'Yes'],
        ['2024-01-22', 'Shell', 'Auto - Gas', 'Fuel for business trip', 60.00, 0.75, '=E5*F5', 'Debit Card', 'Yes'],
        ['2024-02-01', 'Zoom', 'Software/SaaS', 'Video conferencing subscription', 14.99, 1.00, '=E6*F6', 'Credit Card', 'Yes'],
    ]

    for row_data in sample_data:
        sheet.append(row_data)

    # Format data rows
    for row_num in range(2, len(sample_data) + 2):
        # Date format
        sheet[f'A{row_num}'].number_format = 'YYYY-MM-DD'
        sheet[f'A{row_num}'].alignment = Alignment(horizontal='center')

        # Currency format
        for col in ['E', 'G']:
            sheet[f'{col}{row_num}'].number_format = '$#,##0.00'
            sheet[f'{col}{row_num}'].alignment = Alignment(horizontal='right')

        # Percentage format
        sheet[f'F{row_num}'].number_format = '0%'
        sheet[f'F{row_num}'].alignment = Alignment(horizontal='center')

        # Color coding
        sheet[f'E{row_num}'].font = Font(color='0000FF')  # Blue for inputs
        sheet[f'F{row_num}'].font = Font(color='0000FF')  # Blue for inputs
        sheet[f'G{row_num}'].font = Font(color='000000')  # Black for formulas

    # Summary section
    summary_row = len(sample_data) + 3
    sheet[f'B{summary_row}'] = 'EXPENSE SUMMARY BY CATEGORY'
    sheet[f'B{summary_row}'].font = Font(bold=True, size=12)

    # Categories for Schedule C
    categories = [
        'Advertising',
        'Auto - Gas',
        'Auto - Maintenance',
        'Insurance',
        'Legal/Professional',
        'Meals - Business',
        'Office Supplies',
        'Rent/Lease',
        'Software/SaaS',
        'Travel',
        'Utilities',
    ]

    summary_row += 2
    sheet[f'B{summary_row}'] = 'Category'
    sheet[f'D{summary_row}'] = 'Total Deductible'
    sheet[f'B{summary_row}'].font = Font(bold=True)
    sheet[f'D{summary_row}'].font = Font(bold=True)

    for category in categories:
        summary_row += 1
        sheet[f'B{summary_row}'] = category
        sheet[f'D{summary_row}'] = f'=SUMIF(C:C,"{category}",G:G)'
        sheet[f'D{summary_row}'].number_format = '$#,##0.00'

    summary_row += 1
    sheet[f'B{summary_row}'] = 'TOTAL DEDUCTIBLE EXPENSES'
    sheet[f'D{summary_row}'] = '=SUM(G:G)'
    sheet[f'B{summary_row}'].font = Font(bold=True, color='FFFFFF')
    sheet[f'D{summary_row}'].font = Font(bold=True, color='FFFFFF')
    sheet[f'B{summary_row}'].fill = PatternFill('solid', start_color='C65911')
    sheet[f'D{summary_row}'].fill = PatternFill('solid', start_color='C65911')
    sheet[f'D{summary_row}'].number_format = '$#,##0.00'

    # Instructions sheet
    instructions = wb.create_sheet('Instructions')
    instructions['A1'] = 'Expense Tracker Instructions'
    instructions['A1'].font = Font(bold=True, size=14)

    instructions_text = [
        '',
        'How to Use:',
        '1. Enter each business expense on a new row',
        '2. Choose appropriate IRS category from the list',
        '3. Amount is the total expense (Blue = input)',
        '4. Business % is the portion used for business (100% = 1.00, 50% = 0.50)',
        '5. Deductible amount calculates automatically (Black = formula)',
        '6. Keep receipts and note receipt status',
        '',
        'Common IRS Categories (Schedule C):',
        '- Advertising: Marketing, ads, promotional materials',
        '- Auto: Gas, maintenance (track separately for mileage vs actual)',
        '- Insurance: Business liability, E&O insurance',
        '- Legal/Professional: Attorney, CPA, consultants',
        '- Meals - Business: Client meals (50% deductible)',
        '- Office Supplies: Stationery, supplies, small equipment',
        '- Rent/Lease: Office space, equipment rental',
        '- Software/SaaS: Business software subscriptions',
        '- Travel: Airfare, hotels, conference fees',
        '- Utilities: Phone, internet (business portion)',
        '',
        'Business Percentage:',
        '- 100% business use: Enter 1.00',
        '- Mixed use (e.g., home internet): Calculate business portion',
        '- Meals with clients: Automatically 50% deductible per IRS',
        '',
        'Tax Compliance:',
        '- Save all receipts for items over $75',
        '- Document business purpose',
        '- This is for record-keeping only, not tax advice',
        '- Consult a CPA for specific tax questions',
        '',
        f'Created: {datetime.now().strftime("%Y-%m-%d %H:%M")}',
    ]

    for i, text in enumerate(instructions_text, start=2):
        instructions[f'A{i}'] = text
        if text.startswith('-'):
            instructions[f'A{i}'].alignment = Alignment(indent=2)

    # Save workbook
    wb.save(filename)
    print(f"✓ Expense tracker created: {filename}")
    print(f"  - {len(sample_data)} sample transactions")
    print(f"  - {len(categories)} expense categories")
    print(f"  - Summary by category included")
    print(f"  - Instructions sheet added")

if __name__ == '__main__':
    create_expense_tracker()
