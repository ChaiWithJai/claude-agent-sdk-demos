#!/usr/bin/env python3
"""
Income Tracker Template Generator
Creates a spreadsheet for tracking W2 wages and LLC business income.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_income_tracker(filename='Income_Tracker.xlsx'):
    """Create an income tracking spreadsheet with formulas and formatting."""

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Income Tracker"

    # Headers
    headers = ['Date', 'Description', 'Income Type', 'Gross Amount',
               'Withholding', 'Net Amount', 'YTD Total', 'Notes']
    sheet.append(headers)

    # Header formatting
    header_fill = PatternFill('solid', start_color='4472C4')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Column widths
    sheet.column_dimensions['A'].width = 12  # Date
    sheet.column_dimensions['B'].width = 25  # Description
    sheet.column_dimensions['C'].width = 15  # Income Type
    sheet.column_dimensions['D'].width = 14  # Gross Amount
    sheet.column_dimensions['E'].width = 14  # Withholding
    sheet.column_dimensions['F'].width = 14  # Net Amount
    sheet.column_dimensions['G'].width = 14  # YTD Total
    sheet.column_dimensions['H'].width = 30  # Notes

    # Sample data rows
    sample_data = [
        ['2024-01-15', 'January Salary', 'W2 - Regular', 5000, 1200, '=D2-E2', '=F2', 'Bi-weekly paycheck'],
        ['2024-01-31', 'Client ABC Invoice', 'LLC - Consulting', 3500, 0, '=D3-E3', '=G2+F3', 'Project completion'],
        ['2024-02-01', 'February Salary', 'W2 - Regular', 5000, 1200, '=D4-E4', '=G3+F4', 'Bi-weekly paycheck'],
    ]

    for row_data in sample_data:
        sheet.append(row_data)

    # Format data rows
    for row_num in range(2, len(sample_data) + 2):
        # Date format
        sheet[f'A{row_num}'].number_format = 'YYYY-MM-DD'
        sheet[f'A{row_num}'].alignment = Alignment(horizontal='center')

        # Currency format for amounts
        for col in ['D', 'E', 'F', 'G']:
            sheet[f'{col}{row_num}'].number_format = '$#,##0.00'
            sheet[f'{col}{row_num}'].alignment = Alignment(horizontal='right')

        # Color coding
        sheet[f'D{row_num}'].font = Font(color='0000FF')  # Blue for inputs
        sheet[f'E{row_num}'].font = Font(color='0000FF')  # Blue for inputs
        sheet[f'F{row_num}'].font = Font(color='000000')  # Black for formulas
        sheet[f'G{row_num}'].font = Font(color='000000')  # Black for formulas

    # Summary section
    summary_row = len(sample_data) + 3
    sheet[f'B{summary_row}'] = 'SUMMARY'
    sheet[f'B{summary_row}'].font = Font(bold=True, size=12)

    summary_row += 1
    sheet[f'B{summary_row}'] = 'Total W2 Income:'
    sheet[f'D{summary_row}'] = f'=SUMIF(C:C,"W2*",D:D)'
    sheet[f'D{summary_row}'].number_format = '$#,##0.00'
    sheet[f'D{summary_row}'].font = Font(bold=True)

    summary_row += 1
    sheet[f'B{summary_row}'] = 'Total LLC Income:'
    sheet[f'D{summary_row}'] = f'=SUMIF(C:C,"LLC*",D:D)'
    sheet[f'D{summary_row}'].number_format = '$#,##0.00'
    sheet[f'D{summary_row}'].font = Font(bold=True)

    summary_row += 1
    sheet[f'B{summary_row}'] = 'Total Gross Income:'
    sheet[f'D{summary_row}'] = f'=SUM(D2:D100)'
    sheet[f'D{summary_row}'].number_format = '$#,##0.00'
    sheet[f'D{summary_row}'].font = Font(bold=True, color='FFFFFF')
    sheet[f'D{summary_row}'].fill = PatternFill('solid', start_color='4472C4')

    summary_row += 1
    sheet[f'B{summary_row}'] = 'Total Withholding:'
    sheet[f'D{summary_row}'] = f'=SUM(E2:E100)'
    sheet[f'D{summary_row}'].number_format = '$#,##0.00'
    sheet[f'D{summary_row}'].font = Font(bold=True)

    summary_row += 1
    sheet[f'B{summary_row}'] = 'Total Net Income:'
    sheet[f'D{summary_row}'] = f'=SUM(F2:F100)'
    sheet[f'D{summary_row}'].number_format = '$#,##0.00'
    sheet[f'D{summary_row}'].font = Font(bold=True, color='FFFFFF')
    sheet[f'D{summary_row}'].fill = PatternFill('solid', start_color='70AD47')

    # Instructions sheet
    instructions = wb.create_sheet('Instructions')
    instructions['A1'] = 'Income Tracker Instructions'
    instructions['A1'].font = Font(bold=True, size=14)

    instructions_text = [
        '',
        'How to Use:',
        '1. Enter each income transaction on a new row',
        '2. Use consistent Income Type labels (W2 - Regular, LLC - Consulting, etc.)',
        '3. Gross Amount is the total before any deductions (Blue = input)',
        '4. Withholding is taxes taken out (for W2 income)',
        '5. Net Amount and YTD Total calculate automatically (Black = formula)',
        '6. Review the Summary section for totals by income type',
        '',
        'Income Type Categories:',
        '- W2 - Regular: Employment wages',
        '- W2 - Bonus: Employment bonuses',
        '- W2 - Other: Other W2 income',
        '- LLC - Consulting: Business consulting income',
        '- LLC - Services: Business service income',
        '- LLC - Products: Business product sales',
        '- 1099 - Freelance: 1099-NEC income',
        '',
        'Tax Compliance:',
        '- This is for record-keeping only, not tax advice',
        '- Consult a CPA for tax planning',
        '- Keep supporting documentation (pay stubs, invoices)',
        '',
        f'Created: {datetime.now().strftime("%Y-%m-%d %H:%M")}',
    ]

    for i, text in enumerate(instructions_text, start=2):
        instructions[f'A{i}'] = text
        if text.startswith('-'):
            instructions[f'A{i}'].alignment = Alignment(indent=2)

    # Save workbook
    wb.save(filename)
    print(f"✓ Income tracker created: {filename}")
    print(f"  - {len(sample_data)} sample transactions")
    print(f"  - Summary calculations included")
    print(f"  - Instructions sheet added")

if __name__ == '__main__':
    create_income_tracker()
