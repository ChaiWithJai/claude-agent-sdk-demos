#!/usr/bin/env python3
"""
Mileage Log Template Generator
Creates a spreadsheet for tracking business mileage for tax deductions.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_mileage_log(filename='Mileage_Log.xlsx'):
    """Create a mileage tracking log for business travel deductions."""

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Mileage Log 2024"

    # Title
    sheet['A1'] = 'BUSINESS MILEAGE LOG - 2024'
    sheet['A1'].font = Font(bold=True, size=14)
    sheet.merge_cells('A1:H1')

    # IRS Standard Mileage Rate
    sheet['A2'] = '2024 IRS Standard Mileage Rate: $0.67 per mile'
    sheet['A2'].font = Font(italic=True, size=10)
    sheet.merge_cells('A2:H2')

    # Headers
    headers = ['Date', 'Starting Location', 'Destination', 'Business Purpose',
               'Odometer Start', 'Odometer End', 'Miles', 'Deduction']
    row = 4
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill('solid', start_color='203864')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Column widths
    sheet.column_dimensions['A'].width = 12  # Date
    sheet.column_dimensions['B'].width = 25  # Starting Location
    sheet.column_dimensions['C'].width = 25  # Destination
    sheet.column_dimensions['D'].width = 35  # Business Purpose
    sheet.column_dimensions['E'].width = 14  # Odometer Start
    sheet.column_dimensions['F'].width = 14  # Odometer End
    sheet.column_dimensions['G'].width = 10  # Miles
    sheet.column_dimensions['H'].width = 12  # Deduction

    # Sample data
    sample_data = [
        ['2024-01-08', 'Home Office', 'Client ABC Office', 'Sales meeting with Client ABC', 45200, 45235, '=F5-E5', '=G5*0.67'],
        ['2024-01-15', 'Home Office', 'Downtown Conference Center', 'Industry conference attendance', 45235, 45268, '=F6-E6', '=G6*0.67'],
        ['2024-01-22', 'Home Office', 'Office Supply Store', 'Purchase business supplies', 45268, 45280, '=F7-E7', '=G7*0.67'],
        ['2024-02-05', 'Home Office', 'Client XYZ Office', 'Project kickoff meeting', 45280, 45322, '=F8-E8', '=G8*0.67'],
        ['2024-02-12', 'Home Office', 'Airport', 'Business trip to vendor facility', 45322, 45345, '=F9-E9', '=G9*0.67'],
    ]

    for row_data in sample_data:
        sheet.append(row_data)

    # Format data rows
    for row_num in range(5, 5 + len(sample_data)):
        # Date format
        sheet[f'A{row_num}'].number_format = 'YYYY-MM-DD'
        sheet[f'A{row_num}'].alignment = Alignment(horizontal='center')

        # Number format for odometer
        for col in ['E', 'F']:
            sheet[f'{col}{row_num}'].number_format = '#,##0'
            sheet[f'{col}{row_num}'].alignment = Alignment(horizontal='right')
            sheet[f'{col}{row_num}'].font = Font(color='0000FF')  # Blue for inputs

        # Number format for miles
        sheet[f'G{row_num}'].number_format = '#,##0'
        sheet[f'G{row_num}'].alignment = Alignment(horizontal='center')
        sheet[f'G{row_num}'].font = Font(color='000000')  # Black for formulas

        # Currency format for deduction
        sheet[f'H{row_num}'].number_format = '$#,##0.00'
        sheet[f'H{row_num}'].alignment = Alignment(horizontal='right')
        sheet[f'H{row_num}'].font = Font(color='000000')  # Black for formulas

    # Summary section
    summary_row = 5 + len(sample_data) + 2
    sheet[f'A{summary_row}'] = 'SUMMARY'
    sheet[f'A{summary_row}'].font = Font(bold=True, size=12)

    summary_row += 1
    sheet[f'B{summary_row}'] = 'Total Business Miles:'
    sheet[f'C{summary_row}'] = '=SUM(G5:G100)'
    sheet[f'B{summary_row}'].font = Font(bold=True)
    sheet[f'C{summary_row}'].font = Font(bold=True)
    sheet[f'C{summary_row}'].number_format = '#,##0'

    summary_row += 1
    sheet[f'B{summary_row}'] = 'Total Deduction (@ $0.67/mile):'
    sheet[f'C{summary_row}'] = '=SUM(H5:H100)'
    sheet[f'B{summary_row}'].font = Font(bold=True)
    sheet[f'C{summary_row}'].font = Font(bold=True, color='FFFFFF')
    sheet[f'C{summary_row}'].fill = PatternFill('solid', start_color='70AD47')
    sheet[f'C{summary_row}'].number_format = '$#,##0.00'

    summary_row += 2
    sheet[f'B{summary_row}'] = 'NOTE: Record all business trips promptly'
    sheet[f'B{summary_row}'].font = Font(italic=True, size=9)

    # Annual totals section
    annual_row = summary_row + 3
    sheet[f'A{annual_row}'] = 'ANNUAL VEHICLE INFORMATION'
    sheet[f'A{annual_row}'].font = Font(bold=True, size=12)
    sheet.merge_cells(f'A{annual_row}:C{annual_row}')

    annual_row += 1
    sheet[f'B{annual_row}'] = 'Total Miles Driven (All Purposes):'
    sheet[f'C{annual_row}'] = 15000
    sheet[f'C{annual_row}'].number_format = '#,##0'
    sheet[f'C{annual_row}'].font = Font(color='0000FF')

    annual_row += 1
    sheet[f'B{annual_row}'] = 'Business Miles:'
    sheet[f'C{annual_row}'] = '=SUM(G5:G100)'
    sheet[f'C{annual_row}'].number_format = '#,##0'
    sheet[f'C{annual_row}'].font = Font(bold=True)

    annual_row += 1
    sheet[f'B{annual_row}'] = 'Business Use Percentage:'
    sheet[f'C{annual_row}'] = f'=C{annual_row-1}/C{annual_row-2}'
    sheet[f'C{annual_row}'].number_format = '0.0%'
    sheet[f'C{annual_row}'].font = Font(bold=True)

    annual_row += 2
    sheet[f'B{annual_row}'] = 'Vehicle: Make, Model, Year'
    sheet[f'C{annual_row}'] = 'Honda Accord 2020'
    sheet[f'C{annual_row}'].font = Font(color='0000FF')

    annual_row += 1
    sheet[f'B{annual_row}'] = 'Date Placed in Service:'
    sheet[f'C{annual_row}'] = '2020-06-01'
    sheet[f'C{annual_row}'].font = Font(color='0000FF')

    # Instructions sheet
    instructions = wb.create_sheet('Instructions')
    instructions['A1'] = 'Mileage Log Instructions'
    instructions['A1'].font = Font(bold=True, size=14)

    instructions_text = [
        '',
        'IRS Requirements for Mileage Logs:',
        '',
        'The IRS requires detailed records for vehicle deductions. You must track:',
        '  1. Date of each trip',
        '  2. Starting point and destination',
        '  3. Business purpose of the trip',
        '  4. Miles driven (odometer readings or trip distance)',
        '',
        'Two Methods for Vehicle Deductions:',
        '',
        '1. Standard Mileage Rate (Recommended for most):',
        '   - 2024 Rate: $0.67 per business mile',
        '   - Simpler record keeping',
        '   - Track total annual miles and business miles',
        '   - Calculate business use percentage',
        '',
        '2. Actual Expense Method:',
        '   - Track ALL vehicle expenses (gas, maintenance, insurance, depreciation)',
        '   - Apply business use percentage to total expenses',
        '   - More complex but may yield higher deduction',
        '   - Once you use actual expenses, you cannot switch to standard mileage',
        '',
        'How to Use This Log:',
        '',
        '1. Record Every Business Trip:',
        '   - Enter date, locations, and business purpose',
        '   - Record odometer readings at start and end',
        '   - Miles and deduction calculate automatically',
        '',
        '2. What Counts as Business Miles:',
        '   ✓ Travel to meet clients or customers',
        '   ✓ Travel to business meetings',
        '   ✓ Travel to work-related events/conferences',
        '   ✓ Travel to pick up business supplies',
        '   ✓ Travel between work locations',
        '   ✗ Commuting from home to regular workplace',
        '   ✗ Personal errands',
        '',
        '3. Track Annual Vehicle Information:',
        '   - Total miles driven (business + personal)',
        '   - Business use percentage',
        '   - Vehicle details and date placed in service',
        '',
        '4. Record Keeping:',
        '   - Update log promptly after each trip',
        '   - Keep for at least 3 years after filing taxes',
        '   - Can use apps or GPS tracking as backup',
        '',
        'Common Business Trip Examples:',
        '   - Home office to client meeting',
        '   - Home office to office supply store',
        '   - Office to bank for business deposit',
        '   - Between client sites',
        '   - To industry conferences or training',
        '',
        'Tips:',
        '   - Use a smartphone app to track trips in real-time',
        '   - Keep a small notebook in your car',
        '   - Update your log weekly, not at tax time',
        '   - Be specific about business purpose',
        '   - Round-trip miles count for the full trip',
        '',
        '2024 IRS Standard Mileage Rates:',
        '   - Business: $0.67 per mile',
        '   - Medical/Moving: $0.21 per mile',
        '   - Charitable: $0.14 per mile',
        '',
        'DISCLAIMER:',
        '   This template is for record-keeping purposes only and does not',
        '   constitute professional tax advice. Consult with a licensed CPA',
        '   or tax professional for advice specific to your situation.',
        '',
        f'Created: {datetime.now().strftime("%Y-%m-%d %H:%M")}',
    ]

    for i, text in enumerate(instructions_text, start=2):
        instructions[f'A{i}'] = text
        if text.strip().startswith('-') or text.strip().startswith('•'):
            instructions[f'A{i}'].alignment = Alignment(indent=2)
        elif text.strip().startswith('✓') or text.strip().startswith('✗'):
            instructions[f'A{i}'].alignment = Alignment(indent=3)

    # Save workbook
    wb.save(filename)
    print(f"✓ Mileage log created: {filename}")
    print(f"  - {len(sample_data)} sample trips")
    print(f"  - 2024 IRS rate: $0.67/mile")
    print(f"  - Automatic deduction calculations")
    print(f"  - Annual summary section")
    print(f"  - Instructions sheet added")

if __name__ == '__main__':
    create_mileage_log()
