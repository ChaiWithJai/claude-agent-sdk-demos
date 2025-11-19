"""
Deep Benchmarker - Generate normalized KPI comparisons across peers.

Target Users: Private Equity & Strategy Consultants
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from daloopa_client import DaloopaMCPClient, calculate_date_range


def generate_normalized_kpi_comps(
    output_path: str,
    tickers: list[str],
    kpi_name: str,
    num_quarters: int,
    mcp_client: DaloopaMCPClient
) -> dict:
    """
    Generate normalized KPI comparison across multiple tickers.

    Args:
        output_path: Path for output Excel file
        tickers: List of ticker symbols to compare
        kpi_name: KPI name to analyze
        num_quarters: Number of historical quarters
        mcp_client: Daloopa MCP client instance

    Returns:
        dict with comparison summary
    """
    # Calculate date range
    start_date, end_date = calculate_date_range(num_quarters)

    # Collect raw data from all tickers
    all_data = []

    for ticker in tickers:
        response = mcp_client.get_kpi_time_series(
            ticker=ticker,
            kpi_name=kpi_name,
            start_date=start_date,
            end_date=end_date
        )

        for period in response.get("periods", []):
            all_data.append({
                "ticker": ticker,
                "period": period["period"],
                "fiscal_end": period["fiscal_end"],
                "value": period["value"]
            })

    # Normalize fiscal periods to calendar quarters
    normalized = mcp_client.normalize_fiscal_periods(
        data_array=all_data,
        target_calendar="CQ"
    )

    # Build comparison DataFrame
    periods = []
    for np in normalized["normalized_periods"]:
        row = {"Calendar Quarter": np["calendar_period"]}
        for ticker in tickers:
            if ticker in np["tickers"]:
                row[ticker] = np["tickers"][ticker]["value"]
            else:
                row[ticker] = None
        periods.append(row)

    df = pd.DataFrame(periods)
    df = df.sort_values("Calendar Quarter")

    # Calculate statistics
    stats = {
        "Mean": df[tickers].mean(),
        "Median": df[tickers].median(),
        "Min": df[tickers].min(),
        "Max": df[tickers].max(),
        "Std Dev": df[tickers].std()
    }
    stats_df = pd.DataFrame(stats).T

    # Create Excel workbook
    wb = Workbook()

    # Main comparison sheet
    ws_data = wb.active
    ws_data.title = "KPI Comparison"

    # Add title
    ws_data['A1'] = f"{kpi_name} - Normalized Comparison"
    ws_data['A1'].font = Font(bold=True, size=14)
    ws_data.merge_cells(f'A1:{chr(65 + len(tickers))}1')

    # Add metadata
    ws_data['A2'] = f"Tickers: {', '.join(tickers)} | Quarters: {num_quarters}"
    ws_data['A2'].font = Font(italic=True, size=9)

    # Add data
    start_row = 4
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == start_row:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill('solid', start_color='D9E1F2')
                cell.alignment = Alignment(horizontal='center')

    # Format numbers
    is_percentage = "%" in kpi_name.lower()
    for row in range(start_row + 1, ws_data.max_row + 1):
        for col in range(2, len(tickers) + 2):
            cell = ws_data.cell(row=row, column=col)
            if is_percentage:
                cell.number_format = '0.0%'
            else:
                cell.number_format = '#,##0.0'

    # Add statistics section
    stats_start = ws_data.max_row + 3
    ws_data.cell(row=stats_start, column=1, value="Statistics").font = Font(bold=True)

    for r_idx, (stat_name, row) in enumerate(stats_df.iterrows(), stats_start + 1):
        ws_data.cell(row=r_idx, column=1, value=stat_name)
        for c_idx, ticker in enumerate(tickers, 2):
            cell = ws_data.cell(row=r_idx, column=c_idx, value=row[ticker])
            if is_percentage:
                cell.number_format = '0.0%'
            else:
                cell.number_format = '#,##0.0'

    # Create line chart
    chart = LineChart()
    chart.title = f"{kpi_name} Trend"
    chart.style = 10
    chart.y_axis.title = kpi_name
    chart.x_axis.title = "Calendar Quarter"
    chart.height = 10
    chart.width = 15

    # Add data series
    data = Reference(ws_data, min_col=2, max_col=len(tickers)+1,
                     min_row=start_row, max_row=start_row + len(df))
    cats = Reference(ws_data, min_col=1, min_row=start_row+1,
                     max_row=start_row + len(df))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Position chart
    chart_position = f"A{ws_data.max_row + 3}"
    ws_data.add_chart(chart, chart_position)

    # Adjust column widths
    ws_data.column_dimensions['A'].width = 18
    for i in range(len(tickers)):
        ws_data.column_dimensions[chr(66 + i)].width = 12

    # Add rankings sheet
    ws_rank = wb.create_sheet("Rankings")
    create_rankings_sheet(ws_rank, df, tickers, kpi_name)

    wb.save(output_path)

    return {
        "output_file": output_path,
        "kpi": kpi_name,
        "tickers": tickers,
        "quarters_analyzed": len(df),
        "statistics": stats_df.to_dict()
    }


def create_rankings_sheet(ws, df: pd.DataFrame, tickers: list[str], kpi_name: str):
    """Create a sheet showing peer rankings by most recent quarter."""
    ws['A1'] = f"{kpi_name} - Peer Rankings"
    ws['A1'].font = Font(bold=True, size=12)

    # Get most recent quarter data
    latest = df.iloc[-1] if not df.empty else {}

    # Create rankings
    rankings = []
    for ticker in tickers:
        if ticker in latest and pd.notna(latest[ticker]):
            rankings.append((ticker, latest[ticker]))

    rankings.sort(key=lambda x: x[1], reverse=True)

    # Headers
    headers = ["Rank", "Ticker", "Value", "vs. Median"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='D9E1F2')

    # Data
    if rankings:
        median_val = pd.Series([r[1] for r in rankings]).median()

        for rank, (ticker, value) in enumerate(rankings, 1):
            ws.cell(row=rank + 3, column=1, value=rank)
            ws.cell(row=rank + 3, column=2, value=ticker)
            ws.cell(row=rank + 3, column=3, value=value)
            ws.cell(row=rank + 3, column=3).number_format = '#,##0.0'

            vs_median = ((value - median_val) / median_val * 100) if median_val else 0
            cell = ws.cell(row=rank + 3, column=4, value=vs_median / 100)
            cell.number_format = '+0.0%;-0.0%'

            # Highlight top performer
            if rank == 1:
                for col in range(1, 5):
                    ws.cell(row=rank + 3, column=col).fill = PatternFill(
                        'solid', start_color='C6EFCE'
                    )

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12


def auto_discover_peers(
    ticker: str,
    kpi_name: str,
    num_quarters: int,
    output_path: str,
    mcp_client: DaloopaMCPClient
) -> dict:
    """
    Automatically discover peers and generate comparison.

    Args:
        ticker: Primary ticker to find peers for
        kpi_name: KPI to compare
        num_quarters: Number of quarters
        output_path: Output file path
        mcp_client: Daloopa MCP client

    Returns:
        dict with comparison results
    """
    # Get peers from Daloopa
    peer_response = mcp_client.get_peers(ticker, peer_type="direct")

    # Extract ticker list
    peer_tickers = [p["ticker"] for p in peer_response.get("peers", [])]
    all_tickers = [ticker] + peer_tickers[:4]  # Primary + top 4 peers

    return generate_normalized_kpi_comps(
        output_path=output_path,
        tickers=all_tickers,
        kpi_name=kpi_name,
        num_quarters=num_quarters,
        mcp_client=mcp_client
    )


if __name__ == "__main__":
    # Example usage
    client = DaloopaMCPClient()

    result = generate_normalized_kpi_comps(
        output_path="benchmark_output.xlsx",
        tickers=["CRWD", "S", "PANW", "FTNT"],
        kpi_name="Sales and Marketing % of Revenue",
        num_quarters=12,
        mcp_client=client
    )

    print(f"Generated comparison for {len(result['tickers'])} tickers")
    print(f"Analyzed {result['quarters_analyzed']} quarters")
