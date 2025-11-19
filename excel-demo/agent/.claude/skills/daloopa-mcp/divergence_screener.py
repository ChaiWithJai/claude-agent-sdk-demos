"""
Divergence Screener - Find operational KPI divergences from stock performance.

Target Users: L/S Equity & Event-Driven Funds
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from daloopa_client import DaloopaMCPClient


def screen_operational_divergence(
    output_path: str,
    sector: str,
    kpi_name: str,
    kpi_condition: dict,
    stock_condition: dict,
    mcp_client: DaloopaMCPClient
) -> dict:
    """
    Screen for divergences between operational KPIs and stock performance.

    Args:
        output_path: Path for output Excel file
        sector: Daloopa sector name (e.g., "Tier 1 SaaS")
        kpi_name: KPI to screen (e.g., "Net Revenue Retention")
        kpi_condition: {"type": "consecutive_increase", "periods": 2}
        stock_condition: {"metric": "ytd_return", "operator": "<", "value": -10}
        mcp_client: Daloopa MCP client instance

    Returns:
        dict with screening results
    """
    # Screen companies by KPI criteria
    screen_results = mcp_client.screen_kpi(
        sector=sector,
        kpi_name=kpi_name,
        condition=kpi_condition
    )

    # Filter by stock performance
    divergence_matches = []

    for company in screen_results.get("matches", []):
        stock_perf = company["stock_performance_ytd"]

        # Check stock condition
        if evaluate_condition(stock_perf, stock_condition):
            # Calculate divergence score
            kpi_values = company["kpi_values"]
            kpi_change = calculate_kpi_trend(kpi_values)

            divergence_score = abs(kpi_change - stock_perf)

            divergence_matches.append({
                "ticker": company["ticker"],
                "company_name": company["company_name"],
                "kpi_current": kpi_values[-1]["value"] if kpi_values else None,
                "kpi_change_pct": kpi_change,
                "stock_ytd_pct": stock_perf,
                "divergence_score": divergence_score,
                "kpi_history": kpi_values
            })

    # Sort by divergence score
    divergence_matches.sort(key=lambda x: x["divergence_score"], reverse=True)

    # Create Excel output
    wb = Workbook()
    ws = wb.active
    ws.title = "Divergence Screen"

    # Title section
    ws['A1'] = f"Operational Divergence Screen: {kpi_name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')

    condition_desc = f"{kpi_condition['type'].replace('_', ' ').title()} for {kpi_condition.get('periods', 2)} periods"
    ws['A2'] = f"Sector: {sector} | KPI Condition: {condition_desc}"
    ws['A2'].font = Font(italic=True, size=9)
    ws.merge_cells('A2:F2')

    ws['A3'] = f"Stock Filter: YTD {stock_condition['operator']} {stock_condition['value']}%"
    ws['A3'].font = Font(italic=True, size=9)

    # Headers
    headers = ["Ticker", "Company", f"{kpi_name}", "KPI Change %", "Stock YTD %", "Divergence Score"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='D9E1F2')
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, match in enumerate(divergence_matches, 6):
        ws.cell(row=row_idx, column=1, value=match["ticker"])
        ws.cell(row=row_idx, column=2, value=match["company_name"])

        # KPI current value
        kpi_cell = ws.cell(row=row_idx, column=3, value=match["kpi_current"])
        if match["kpi_current"] and match["kpi_current"] < 10:
            kpi_cell.number_format = '0.0%'
        else:
            kpi_cell.number_format = '#,##0.0'

        # KPI change
        ws.cell(row=row_idx, column=4, value=match["kpi_change_pct"] / 100)
        ws.cell(row=row_idx, column=4).number_format = '+0.0%;-0.0%'

        # Color code KPI trend
        if match["kpi_change_pct"] > 0:
            ws.cell(row=row_idx, column=4).font = Font(color="006400")
        else:
            ws.cell(row=row_idx, column=4).font = Font(color="8B0000")

        # Stock YTD
        ws.cell(row=row_idx, column=5, value=match["stock_ytd_pct"] / 100)
        ws.cell(row=row_idx, column=5).number_format = '+0.0%;-0.0%'

        if match["stock_ytd_pct"] < 0:
            ws.cell(row=row_idx, column=5).font = Font(color="8B0000")

        # Divergence score
        ws.cell(row=row_idx, column=6, value=match["divergence_score"])
        ws.cell(row=row_idx, column=6).number_format = '0.0'

        # Highlight high divergence opportunities
        if match["divergence_score"] > 20:
            ws.cell(row=row_idx, column=6).fill = PatternFill('solid', start_color='90EE90')

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18

    # Add divergence chart
    if divergence_matches:
        chart = BarChart()
        chart.title = "Divergence Scores"
        chart.type = "col"
        chart.height = 8
        chart.width = 12

        data = Reference(ws, min_col=6, min_row=5,
                        max_row=5 + len(divergence_matches))
        cats = Reference(ws, min_col=1, min_row=6,
                        max_row=5 + len(divergence_matches))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        chart_position = f"H5"
        ws.add_chart(chart, chart_position)

    # Investment thesis section
    thesis_row = max(6 + len(divergence_matches), 20)
    ws.cell(row=thesis_row, column=1, value="Investment Thesis").font = Font(bold=True, size=12)

    if divergence_matches:
        top_match = divergence_matches[0]
        thesis_text = (
            f"Top opportunity: {top_match['ticker']} ({top_match['company_name']}) "
            f"shows {kpi_name} improving (+{top_match['kpi_change_pct']:.1f}%) "
            f"while stock is down {top_match['stock_ytd_pct']:.1f}% YTD. "
            f"This {abs(top_match['divergence_score']):.1f}pt divergence suggests potential mispricing."
        )
        thesis_cell = ws.cell(row=thesis_row + 1, column=1, value=thesis_text)
        ws.merge_cells(f'A{thesis_row + 1}:F{thesis_row + 1}')

        # Add follow-up actions
        ws.cell(row=thesis_row + 3, column=1, value="Suggested Actions:").font = Font(bold=True)
        actions = [
            f"1. Review {top_match['ticker']} recent earnings call for management commentary",
            f"2. Analyze {kpi_name} drivers and sustainability",
            f"3. Check for one-time items affecting stock performance",
            "4. Validate with Daloopa source documents for data accuracy"
        ]
        for i, action in enumerate(actions):
            ws.cell(row=thesis_row + 4 + i, column=1, value=action)
    else:
        ws.cell(row=thesis_row + 1, column=1,
                value="No companies matched the screening criteria.")

    # Add detail sheet for top opportunities
    if divergence_matches:
        ws_detail = wb.create_sheet("Top Opportunities")
        create_detail_sheet(ws_detail, divergence_matches[:5], kpi_name)

    wb.save(output_path)

    return {
        "output_file": output_path,
        "sector": sector,
        "kpi": kpi_name,
        "total_screened": len(screen_results.get("matches", [])),
        "matches_found": len(divergence_matches),
        "top_opportunities": divergence_matches[:5]
    }


def create_detail_sheet(ws, matches: list, kpi_name: str):
    """Create detailed analysis sheet for top opportunities."""
    ws['A1'] = "Top 5 Divergence Opportunities - Detail"
    ws['A1'].font = Font(bold=True, size=12)

    row = 3
    for match in matches:
        # Company header
        ws.cell(row=row, column=1,
                value=f"{match['ticker']} - {match['company_name']}").font = Font(bold=True)
        row += 1

        # KPI history
        ws.cell(row=row, column=1, value=f"{kpi_name} History:")
        for kpi_val in match.get("kpi_history", []):
            ws.cell(row=row, column=2, value=kpi_val["period"])
            ws.cell(row=row, column=3, value=kpi_val["value"])
            row += 1

        # Key metrics
        ws.cell(row=row, column=1, value="KPI Change:")
        ws.cell(row=row, column=2, value=f"{match['kpi_change_pct']:+.1f}%")
        row += 1

        ws.cell(row=row, column=1, value="Stock YTD:")
        ws.cell(row=row, column=2, value=f"{match['stock_ytd_pct']:+.1f}%")
        row += 1

        ws.cell(row=row, column=1, value="Divergence:")
        ws.cell(row=row, column=2, value=f"{match['divergence_score']:.1f}")
        row += 2

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12


def evaluate_condition(value: float, condition: dict) -> bool:
    """Evaluate a condition against a value."""
    operator = condition.get("operator", "<")
    threshold = condition.get("value", 0)

    operators = {
        "<": lambda v, t: v < t,
        ">": lambda v, t: v > t,
        "<=": lambda v, t: v <= t,
        ">=": lambda v, t: v >= t,
        "==": lambda v, t: v == t
    }

    return operators.get(operator, lambda v, t: False)(value, threshold)


def calculate_kpi_trend(kpi_values: list) -> float:
    """Calculate percentage change in KPI over periods."""
    if len(kpi_values) < 2:
        return 0

    first_value = kpi_values[0]["value"]
    last_value = kpi_values[-1]["value"]

    if first_value == 0:
        return 0

    return ((last_value - first_value) / first_value) * 100


if __name__ == "__main__":
    # Example usage
    client = DaloopaMCPClient()

    result = screen_operational_divergence(
        output_path="divergence_screen.xlsx",
        sector="Tier 1 SaaS",
        kpi_name="Net Revenue Retention",
        kpi_condition={"type": "consecutive_increase", "periods": 2},
        stock_condition={"metric": "ytd_return", "operator": "<", "value": -10},
        mcp_client=client
    )

    print(f"Screened {result['total_screened']} companies")
    print(f"Found {result['matches_found']} divergence opportunities")

    if result['top_opportunities']:
        top = result['top_opportunities'][0]
        print(f"Top pick: {top['ticker']} with {top['divergence_score']:.1f} divergence score")
