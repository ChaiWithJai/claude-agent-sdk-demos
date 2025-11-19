"""
Earnings Sprint Agent - Update models with new actuals and variance analysis.

Target Users: L/S Hedge Funds during earnings season
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.comments import Comment
from daloopa_client import DaloopaMCPClient, get_quarter_dates


def update_model_variance_analysis(
    model_path: str,
    ticker: str,
    kpis: list[str],
    quarter: str,
    estimate_column: str,
    mcp_client: DaloopaMCPClient
) -> dict:
    """
    Update Excel model with Daloopa actuals and calculate variance.

    Args:
        model_path: Path to the Excel model file
        ticker: Stock ticker symbol
        kpis: List of KPI names to update
        quarter: Quarter to fetch (e.g., "Q3 2024")
        estimate_column: Column letter containing estimates
        mcp_client: Daloopa MCP client instance

    Returns:
        dict with update summary and variances
    """
    wb = load_workbook(model_path)
    sheet = wb.active

    results = {"ticker": ticker, "quarter": quarter, "updates": []}

    # Find the column for actuals (next to estimates)
    actual_col = chr(ord(estimate_column) + 1)
    variance_col = chr(ord(estimate_column) + 2)

    # Add headers
    header_row = 1
    sheet[f'{actual_col}{header_row}'] = f'{quarter} Actual'
    sheet[f'{actual_col}{header_row}'].font = Font(bold=True)
    sheet[f'{variance_col}{header_row}'] = f'{quarter} Variance'
    sheet[f'{variance_col}{header_row}'].font = Font(bold=True)

    # Get quarter date range
    start_date, end_date = get_quarter_dates(quarter)

    for kpi in kpis:
        # Fetch actual from Daloopa MCP
        response = mcp_client.get_kpi_time_series(
            ticker=ticker,
            kpi_name=kpi,
            start_date=start_date,
            end_date=end_date
        )

        if not response.get("periods"):
            continue

        actual_data = response["periods"][0]
        actual_value = actual_data["value"]
        data_point_id = actual_data["data_point_id"]

        # Find KPI row in model
        kpi_row = find_kpi_row(sheet, kpi)
        if not kpi_row:
            results["updates"].append({
                "kpi": kpi,
                "error": f"KPI '{kpi}' not found in model"
            })
            continue

        # Get estimate value
        estimate_cell = sheet[f'{estimate_column}{kpi_row}']
        estimate_value = estimate_cell.value or 0

        # Insert actual value (blue for inputs)
        actual_cell = sheet[f'{actual_col}{kpi_row}']
        actual_cell.value = actual_value
        actual_cell.font = Font(color="0000FF")
        actual_cell.number_format = '#,##0.0'

        # Calculate variance formula (black for formulas)
        variance_cell = sheet[f'{variance_col}{kpi_row}']
        variance_cell.value = f'={actual_col}{kpi_row}-{estimate_column}{kpi_row}'
        variance_cell.font = Font(color="000000")
        variance_cell.number_format = '+#,##0.0;-#,##0.0'

        # Get source context and add comment
        source = mcp_client.get_source_context(data_point_id)

        comment_text = (
            f"Source: {source['source_type']}, {source['filing_date']}\n"
            f"Page: {source['page_number']}\n"
            f"URL: {source['pdf_url']}"
        )
        actual_cell.comment = Comment(comment_text, "Daloopa MCP")

        # Calculate variance for summary
        variance = actual_value - estimate_value
        variance_pct = (variance / estimate_value * 100) if estimate_value else 0

        # Highlight significant variances (>5%)
        if abs(variance_pct) > 5:
            variance_cell.fill = PatternFill('solid', start_color='FFFF00')

        results["updates"].append({
            "kpi": kpi,
            "estimate": estimate_value,
            "actual": actual_value,
            "variance": variance,
            "variance_pct": round(variance_pct, 2),
            "source_url": source['pdf_url']
        })

    wb.save(model_path)

    # Summary statistics
    variances = [u["variance_pct"] for u in results["updates"] if "variance_pct" in u]
    if variances:
        results["summary"] = {
            "total_updated": len(variances),
            "avg_variance_pct": round(sum(variances) / len(variances), 2),
            "max_variance_pct": round(max(variances, key=abs), 2),
            "significant_variances": sum(1 for v in variances if abs(v) > 5)
        }

    return results


def find_kpi_row(sheet, kpi_name: str) -> int:
    """Find row number containing KPI name in column A."""
    kpi_lower = kpi_name.lower()

    for row in range(1, sheet.max_row + 1):
        cell_value = sheet[f'A{row}'].value
        if cell_value and kpi_lower in str(cell_value).lower():
            return row

    return None


def batch_update_models(
    model_configs: list[dict],
    mcp_client: DaloopaMCPClient
) -> list[dict]:
    """
    Update multiple models in batch.

    Args:
        model_configs: List of dicts with model_path, ticker, kpis, quarter, estimate_column
        mcp_client: Daloopa MCP client instance

    Returns:
        list: Results for each model
    """
    results = []

    for config in model_configs:
        try:
            result = update_model_variance_analysis(
                model_path=config["model_path"],
                ticker=config["ticker"],
                kpis=config["kpis"],
                quarter=config["quarter"],
                estimate_column=config.get("estimate_column", "B"),
                mcp_client=mcp_client
            )
            results.append(result)
        except Exception as e:
            results.append({
                "ticker": config.get("ticker"),
                "error": str(e)
            })

    return results


if __name__ == "__main__":
    # Example usage
    client = DaloopaMCPClient()

    result = update_model_variance_analysis(
        model_path="sample_model.xlsx",
        ticker="DASH",
        kpis=["Total Orders", "Marketplace GOV", "Revenue"],
        quarter="Q3 2024",
        estimate_column="B",
        mcp_client=client
    )

    print(f"Updated {result['summary']['total_updated']} KPIs")
    print(f"Significant variances: {result['summary']['significant_variances']}")
