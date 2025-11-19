"""
Proxy Model Builder - Apply public company cost structures to private targets.

Target Users: Private Equity / Corporate Dev
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from daloopa_client import DaloopaMCPClient


def build_proxy_cost_structure(
    output_path: str,
    private_revenue: float,
    proxy_ticker: str,
    proxy_year: str,
    revenue_unit: str,
    mcp_client: DaloopaMCPClient
) -> dict:
    """
    Build pro-forma P&L using public company cost structure as proxy.

    Args:
        output_path: Path for output Excel file
        private_revenue: Private company revenue
        proxy_ticker: Public company ticker to use as proxy
        proxy_year: Historical year for proxy margins
        revenue_unit: "millions" or "thousands"
        mcp_client: Daloopa MCP client instance

    Returns:
        dict with P&L summary
    """
    # Define P&L line items to fetch
    cost_structure_kpis = [
        ("Revenue", "Revenue"),
        ("COGS", "Cost of Goods Sold"),
        ("Gross Profit", "Gross Profit"),
        ("R&D", "Research and Development"),
        ("S&M", "Sales and Marketing"),
        ("G&A", "General and Administrative"),
        ("Operating Income", "Operating Income")
    ]

    # Fetch proxy company historical data
    year_start = f"{proxy_year}-01-01"
    year_end = f"{proxy_year}-12-31"

    proxy_data = {}
    source_info = {}

    for kpi_short, kpi_name in cost_structure_kpis:
        response = mcp_client.get_kpi_time_series(
            ticker=proxy_ticker,
            kpi_name=kpi_name,
            start_date=year_start,
            end_date=year_end,
            period_type="annual"
        )

        if response.get("periods"):
            period_data = response["periods"][0]
            proxy_data[kpi_short] = period_data["value"]
            source_info[kpi_short] = period_data.get("data_point_id")

    # Calculate margins (as % of revenue)
    proxy_revenue = proxy_data.get("Revenue", 1)
    margins = {}

    for kpi_short, _ in cost_structure_kpis:
        if kpi_short in proxy_data and kpi_short != "Revenue":
            margins[kpi_short] = proxy_data[kpi_short] / proxy_revenue

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Pro-Forma P&L"

    # Styling
    header_font = Font(bold=True, size=12)
    input_font = Font(color="0000FF")  # Blue for inputs
    formula_font = Font(color="000000")  # Black for formulas
    thin_border = Border(bottom=Side(style='thin'))
    thick_border = Border(bottom=Side(style='medium'))

    # Title
    ws['A1'] = "Pro-Forma P&L - Private Company Analysis"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    ws['A2'] = f"Proxy: {proxy_ticker} ({proxy_year} margins)"
    ws['A2'].font = Font(italic=True, size=10)

    # Input section
    ws['A4'] = "Assumptions"
    ws['A4'].font = header_font

    ws['A5'] = "Private Company Revenue"
    ws['B5'] = private_revenue
    ws['B5'].font = input_font
    ws['B5'].number_format = '$#,##0'
    ws['B5'].fill = PatternFill('solid', start_color='FFFF99')
    ws['C5'] = f"${revenue_unit[0].upper()}"

    ws['A6'] = f"Proxy Revenue ({proxy_ticker} {proxy_year})"
    ws['B6'] = proxy_revenue
    ws['B6'].number_format = '$#,##0'
    ws['C6'] = f"${revenue_unit[0].upper()}"

    # P&L Statement
    ws['A8'] = "Pro-Forma Income Statement"
    ws['A8'].font = header_font

    # Headers
    headers = ["Line Item", "Amount", "% of Rev", f"{proxy_ticker} %", "Variance"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='D9E1F2')
        cell.alignment = Alignment(horizontal='center')

    # Build P&L rows
    row = 10

    # Revenue
    ws.cell(row=row, column=1, value="Revenue")
    ws.cell(row=row, column=2, value="=B5")
    ws.cell(row=row, column=2).font = formula_font
    ws.cell(row=row, column=2).number_format = '$#,##0'
    ws.cell(row=row, column=3, value="=B10/B10")
    ws.cell(row=row, column=3).number_format = '0.0%'
    ws.cell(row=row, column=4, value=1.0)
    ws.cell(row=row, column=4).number_format = '0.0%'
    ws.cell(row=row, column=5, value="=C10-D10")
    ws.cell(row=row, column=5).number_format = '+0.0%;-0.0%'
    row += 1

    # COGS
    cogs_margin = margins.get("COGS", 0.4)
    ws.cell(row=row, column=1, value="Cost of Goods Sold")
    ws.cell(row=row, column=2, value=f"=B$5*D{row}")
    ws.cell(row=row, column=2).font = formula_font
    ws.cell(row=row, column=2).number_format = '($#,##0)'
    ws.cell(row=row, column=3, value=f"=B{row}/B$10")
    ws.cell(row=row, column=3).number_format = '0.0%'
    ws.cell(row=row, column=4, value=cogs_margin)
    ws.cell(row=row, column=4).font = input_font
    ws.cell(row=row, column=4).number_format = '0.0%'
    ws.cell(row=row, column=5, value=f"=C{row}-D{row}")
    ws.cell(row=row, column=5).number_format = '+0.0%;-0.0%'
    row += 1

    # Gross Profit
    gp_row = row
    ws.cell(row=row, column=1, value="Gross Profit")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value="=B10-B11")
    ws.cell(row=row, column=2).font = formula_font
    ws.cell(row=row, column=2).number_format = '$#,##0'
    ws.cell(row=row, column=3, value=f"=B{row}/B$10")
    ws.cell(row=row, column=3).number_format = '0.0%'
    gp_margin = margins.get("Gross Profit", 0.6)
    ws.cell(row=row, column=4, value=gp_margin)
    ws.cell(row=row, column=4).number_format = '0.0%'
    ws.cell(row=row, column=5, value=f"=C{row}-D{row}")
    ws.cell(row=row, column=5).number_format = '+0.0%;-0.0%'
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

    # Blank row
    row += 1

    # Operating expenses header
    ws.cell(row=row, column=1, value="Operating Expenses")
    ws.cell(row=row, column=1).font = Font(italic=True)
    row += 1

    # Operating expense items
    opex_items = [
        ("Research & Development", "R&D", 0.1),
        ("Sales & Marketing", "S&M", 0.15),
        ("General & Administrative", "G&A", 0.08)
    ]

    opex_rows = []
    for item_name, margin_key, default_margin in opex_items:
        margin_val = margins.get(margin_key, default_margin)
        ws.cell(row=row, column=1, value=item_name)
        ws.cell(row=row, column=2, value=f"=B$5*D{row}")
        ws.cell(row=row, column=2).font = formula_font
        ws.cell(row=row, column=2).number_format = '($#,##0)'
        ws.cell(row=row, column=3, value=f"=B{row}/B$10")
        ws.cell(row=row, column=3).number_format = '0.0%'
        ws.cell(row=row, column=4, value=margin_val)
        ws.cell(row=row, column=4).font = input_font
        ws.cell(row=row, column=4).number_format = '0.0%'
        ws.cell(row=row, column=5, value=f"=C{row}-D{row}")
        ws.cell(row=row, column=5).number_format = '+0.0%;-0.0%'
        opex_rows.append(row)
        row += 1

    # Total Operating Expenses
    ws.cell(row=row, column=1, value="Total Operating Expenses")
    opex_sum = "+".join([f"B{r}" for r in opex_rows])
    ws.cell(row=row, column=2, value=f"={opex_sum}")
    ws.cell(row=row, column=2).font = formula_font
    ws.cell(row=row, column=2).number_format = '($#,##0)'
    ws.cell(row=row, column=3, value=f"=B{row}/B$10")
    ws.cell(row=row, column=3).number_format = '0.0%'
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = thin_border
    total_opex_row = row
    row += 1

    # Blank row
    row += 1

    # Operating Income
    ws.cell(row=row, column=1, value="Operating Income (EBIT)")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"=B{gp_row}-B{total_opex_row}")
    ws.cell(row=row, column=2).font = formula_font
    ws.cell(row=row, column=2).number_format = '$#,##0;($#,##0)'
    ws.cell(row=row, column=3, value=f"=B{row}/B$10")
    ws.cell(row=row, column=3).number_format = '0.0%'
    op_margin = margins.get("Operating Income", 0.1)
    ws.cell(row=row, column=4, value=op_margin)
    ws.cell(row=row, column=4).number_format = '0.0%'
    ws.cell(row=row, column=5, value=f"=C{row}-D{row}")
    ws.cell(row=row, column=5).number_format = '+0.0%;-0.0%'
    ws.cell(row=row, column=2).fill = PatternFill('solid', start_color='E8F5E9')
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = thick_border

    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    # Source notes
    notes_row = row + 3
    ws.cell(row=notes_row, column=1, value="Source Notes").font = Font(bold=True)
    ws.cell(row=notes_row + 1, column=1,
            value=f"Proxy margins: {proxy_ticker} {proxy_year} annual report (Daloopa MCP)")
    ws.cell(row=notes_row + 2, column=1,
            value="Blue values = adjustable inputs | Yellow highlight = key assumption")
    ws.cell(row=notes_row + 3, column=1,
            value="Adjust D column margins to test different scenarios")

    # Add sensitivity analysis sheet
    ws_sens = wb.create_sheet("Sensitivity")
    create_sensitivity_analysis(ws_sens, private_revenue, margins)

    # Add margin assumptions sheet
    ws_margins = wb.create_sheet("Proxy Margins")
    create_margin_sheet(ws_margins, proxy_ticker, proxy_year, proxy_data, margins)

    wb.save(output_path)

    return {
        "output_file": output_path,
        "proxy_ticker": proxy_ticker,
        "proxy_year": proxy_year,
        "private_revenue": private_revenue,
        "revenue_unit": revenue_unit,
        "margins_applied": margins,
        "sheets_created": ["Pro-Forma P&L", "Sensitivity", "Proxy Margins"]
    }


def create_sensitivity_analysis(ws, base_revenue: float, margins: dict):
    """Create sensitivity analysis table."""
    ws['A1'] = "Operating Income Sensitivity Analysis"
    ws['A1'].font = Font(bold=True, size=12)

    ws['A3'] = "Revenue Growth vs Gross Margin Change"
    ws['A3'].font = Font(bold=True)

    # Headers
    ws['B4'] = "Gross Margin Change"
    margin_changes = [-5, -2.5, 0, 2.5, 5]
    revenue_changes = [-20, -10, 0, 10, 20]

    for col, mc in enumerate(margin_changes, 3):
        ws.cell(row=4, column=col, value=f"{mc:+.1f}%")
        ws.cell(row=4, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=4, column=col).font = Font(bold=True)

    ws['A5'] = "Revenue Growth"
    ws['A5'].font = Font(bold=True)

    for row, rc in enumerate(revenue_changes, 5):
        ws.cell(row=row, column=2, value=f"{rc:+d}%")
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')

        for col, mc in enumerate(margin_changes, 3):
            adj_rev = base_revenue * (1 + rc/100)
            adj_gm = margins.get("Gross Profit", 0.6) + mc/100
            adj_opex = (margins.get("R&D", 0.1) +
                       margins.get("S&M", 0.15) +
                       margins.get("G&A", 0.08))
            op_income = adj_rev * (adj_gm - adj_opex)

            cell = ws.cell(row=row, column=col, value=op_income)
            cell.number_format = '$#,##0'

            # Color code
            if op_income > 0:
                cell.fill = PatternFill('solid', start_color='C6EFCE')
            else:
                cell.fill = PatternFill('solid', start_color='FFC7CE')

    # Operating margin sensitivity
    ws['A12'] = "Operating Margin % Sensitivity"
    ws['A12'].font = Font(bold=True)

    for row, rc in enumerate(revenue_changes, 13):
        ws.cell(row=row, column=2, value=f"{rc:+d}%")

        for col, mc in enumerate(margin_changes, 3):
            adj_gm = margins.get("Gross Profit", 0.6) + mc/100
            adj_opex = (margins.get("R&D", 0.1) +
                       margins.get("S&M", 0.15) +
                       margins.get("G&A", 0.08))
            op_margin = adj_gm - adj_opex

            cell = ws.cell(row=row, column=col, value=op_margin)
            cell.number_format = '0.0%'

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12


def create_margin_sheet(ws, ticker: str, year: str, proxy_data: dict, margins: dict):
    """Create sheet showing proxy company margin details."""
    ws['A1'] = f"{ticker} {year} Cost Structure"
    ws['A1'].font = Font(bold=True, size=12)

    headers = ["Line Item", "Amount ($M)", "% of Revenue"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='D9E1F2')

    items = [
        ("Revenue", "Revenue"),
        ("Cost of Goods Sold", "COGS"),
        ("Gross Profit", "Gross Profit"),
        ("R&D", "R&D"),
        ("Sales & Marketing", "S&M"),
        ("G&A", "G&A"),
        ("Operating Income", "Operating Income")
    ]

    row = 4
    for display_name, key in items:
        ws.cell(row=row, column=1, value=display_name)
        ws.cell(row=row, column=2, value=proxy_data.get(key, 0))
        ws.cell(row=row, column=2).number_format = '#,##0.0'

        if key == "Revenue":
            ws.cell(row=row, column=3, value=1.0)
        else:
            ws.cell(row=row, column=3, value=margins.get(key, 0))
        ws.cell(row=row, column=3).number_format = '0.0%'

        row += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15


if __name__ == "__main__":
    # Example usage
    client = DaloopaMCPClient()

    result = build_proxy_cost_structure(
        output_path="proxy_model.xlsx",
        private_revenue=50,
        proxy_ticker="ELF",
        proxy_year="2019",
        revenue_unit="millions",
        mcp_client=client
    )

    print(f"Created pro-forma P&L at {result['output_file']}")
    print(f"Proxy: {result['proxy_ticker']} {result['proxy_year']}")
    print(f"Sheets: {', '.join(result['sheets_created'])}")
