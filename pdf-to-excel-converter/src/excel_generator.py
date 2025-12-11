from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from datetime import datetime

class ExcelGenerator:
    """Generate formatted Excel files for smoke detection industry data."""

    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    def create_boq_sheet(self, df: pd.DataFrame, sheet_name: str = "BOQ") -> None:
        """Create a formatted BOQ sheet."""
        ws = self.wb.create_sheet(sheet_name)

        ws['A1'] = 'BILL OF QUANTITIES - SMOKE DETECTION SYSTEM'
        ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:F1')

        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        ws['A2'].font = Font(size=9, italic=True)

        header_row = 4
        headers = ['Item No', 'Description', 'Quantity', 'Unit', 'Category', 'Unit Price', 'Total Price']

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        data_start_row = header_row + 1

        for idx, row in df.iterrows():
            row_num = data_start_row + idx
            ws.cell(row=row_num, column=1, value=row.get('Item No', ''))
            ws.cell(row=row_num, column=2, value=row.get('Description', ''))
            ws.cell(row=row_num, column=3, value=row.get('Quantity', ''))
            ws.cell(row=row_num, column=4, value=row.get('Unit', 'NOS'))
            ws.cell(row=row_num, column=5, value=row.get('Category', 'Other'))

            unit_price_cell = ws.cell(row=row_num, column=6)
            unit_price_cell.value = 0
            unit_price_cell.number_format = '$#,##0.00'
            unit_price_cell.font = Font(color='0000FF')

            total_cell = ws.cell(row=row_num, column=7)
            total_cell.value = f'=C{row_num}*F{row_num}'
            total_cell.number_format = '$#,##0.00'

            for col in range(1, 8):
                cell = ws.cell(row=row_num, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

        last_data_row = data_start_row + len(df)

        summary_row = last_data_row + 2
        ws.cell(row=summary_row, column=6, value='TOTAL:').font = Font(bold=True)
        ws.cell(row=summary_row, column=6).alignment = Alignment(horizontal='right')
        total_formula_cell = ws.cell(row=summary_row, column=7)
        total_formula_cell.value = f'=SUM(G{data_start_row}:G{last_data_row})'
        total_formula_cell.font = Font(bold=True)
        total_formula_cell.number_format = '$#,##0.00'
        total_formula_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15

    def create_spec_sheet(self, df: pd.DataFrame, sheet_name: str = "Specifications") -> None:
        """Create a formatted specification sheet."""
        ws = self.wb.create_sheet(sheet_name)

        ws['A1'] = 'PRODUCT SPECIFICATIONS - SMOKE DETECTION EQUIPMENT'
        ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:C1')

        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        ws['A2'].font = Font(size=9, italic=True)

        header_row = 4
        headers = ['Category', 'Specification', 'Value']

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        data_start_row = header_row + 1
        current_category = None

        for idx, row in df.iterrows():
            row_num = data_start_row + idx
            category = row.get('Category', '')

            if category != current_category:
                ws.cell(row=row_num, column=1, value=category).font = Font(bold=True)
                ws.cell(row=row_num, column=1).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                current_category = category
            else:
                ws.cell(row=row_num, column=1, value='')

            ws.cell(row=row_num, column=2, value=row.get('Specification', ''))
            ws.cell(row=row_num, column=3, value=row.get('Value', ''))

            for col in range(1, 4):
                cell = ws.cell(row=row_num, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40

    def create_summary_sheet(self, boq_df: pd.DataFrame = None) -> None:
        """Create a summary sheet with category breakdown."""
        ws = self.wb.create_sheet("Summary", 0)

        ws['A1'] = 'PROJECT SUMMARY - SMOKE DETECTION SYSTEM'
        ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')

        ws['A3'] = 'Category Breakdown'
        ws['A3'].font = Font(bold=True, size=12)

        if boq_df is not None and 'Category' in boq_df.columns:
            summary = boq_df.groupby('Category').size().reset_index(name='Item Count')

            header_row = 4
            ws.cell(row=header_row, column=1, value='Category').font = Font(bold=True)
            ws.cell(row=header_row, column=2, value='Item Count').font = Font(bold=True)

            for idx, row in summary.iterrows():
                row_num = header_row + 1 + idx
                ws.cell(row=row_num, column=1, value=row['Category'])
                ws.cell(row=row_num, column=2, value=row['Item Count'])

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15

    def save(self, filename: str) -> None:
        """Save the workbook."""
        self.wb.save(filename)
